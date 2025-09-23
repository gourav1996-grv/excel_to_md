import pandas as pd
from openpyxl import load_workbook
import re
from io import BytesIO


def is_likely_header(row_values):
    """
    Heuristic to detect if a row is a header: multiple columns, short words, no ending punctuation.
    """
    str_values = [str(v).strip() for v in row_values if v is not None]
    if len(str_values) < 2:  # Need at least 2 columns for a header
        return False
    avg_len = sum(len(s) for s in str_values) / len(str_values)
    if avg_len > 25:  # Headers are usually short
        return False
    # No sentence-ending punctuation in headers
    if any(s.endswith((".", "?", "!")) for s in str_values):
        return False
    # Common header words or all caps can be added if needed, but keep general
    return True


def extract_sections(sheet):
    """
    Extract text blocks and tables from the sheet.
    """
    sections = []
    row_num = 1
    max_row = sheet.max_row
    while row_num <= max_row:
        # Skip empty rows
        while row_num <= max_row and not any(
            cell.value for cell in sheet[row_num] if cell.value is not None
        ):
            row_num += 1
        if row_num > max_row:
            break

        current_row_values = [
            cell.value for cell in sheet[row_num] if cell.value is not None
        ]

        if is_likely_header(current_row_values):
            # Potential table start
            table_start = row_num
            header_row = current_row_values
            row_num += 1
            # Collect rows until empty or end
            table_rows = [header_row]
            while row_num <= max_row and any(
                cell.value for cell in sheet[row_num] if cell.value is not None
            ):
                row_values = [cell.value for cell in sheet[row_num]]
                # Pad with None if shorter than header
                row_values += [None] * (len(header_row) - len(row_values))
                table_rows.append(row_values)
                row_num += 1
            # Create DF from collected rows
            df = pd.DataFrame(table_rows[1:], columns=table_rows[0])
            df = df.dropna(how="all").dropna(axis=1, how="all")
            # Handle multi-line: replace \n with ' ; '
            df = df.applymap(
                lambda x: str(x).replace("\n", " ; ") if pd.notnull(x) else ""
            )
            table_md = df.to_markdown(index=False)
            sections.append(("table", table_md))
        else:
            # Text block
            text_lines = []
            while (
                row_num <= max_row
                and not is_likely_header(
                    [cell.value for cell in sheet[row_num] if cell.value is not None]
                )
                and any(cell.value for cell in sheet[row_num] if cell.value is not None)
            ):
                row_values = [
                    cell.value for cell in sheet[row_num] if cell.value is not None
                ]
                text_line = " ".join(str(v).strip() for v in row_values)
                if text_line:
                    text_lines.append(text_line)
                row_num += 1
            if text_lines:
                sections.append(("text", "\n".join(text_lines)))

    return sections


def convert_excel_to_markdown(file_path, output_md="data_dict_markdown.md"):
    wb = load_workbook(file_path, data_only=True)
    markdown_docs = []

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        sections = extract_sections(sheet)

        full_md = f"# Sheet: {sheet_name}\n"
        text_count = 1
        table_count = 1
        for sec_type, content in sections:
            if sec_type == "text":
                full_md += f"## Text Section {text_count}:\n{content}\n\n"
                text_count += 1
            elif sec_type == "table":
                full_md += f"## Table Section {table_count}:\n{content}\n\n"
                table_count += 1

        # Clean extra newlines
        full_md = re.sub(r"\n{3,}", "\n\n", full_md)
        markdown_docs.append(full_md)

    with open(output_md, "w", encoding="utf-8") as f:
        f.write("\n\n---\n\n".join(markdown_docs))

    print(f"Conversion complete. Markdown saved to {output_md}")


if __name__ == "__main__":
    convert_excel_to_markdown(
        "D:\work\dummy_data_dictionary.xlsx"
    )  # Replace with your actual file path
